import re
import time

import requests
from bs4 import BeautifulSoup
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook


def main():
    url = 'https://www.w3school.com.cn/python/python_while_loops.asp'  # 要爬取的网页地址
    # url="https://www.baidu.com/"
    # url="https://www.baidu.com/s?wd=%E4%B9%A0%E8%BF%91%E5%B9%B3%E4%B8%BB%E6%8C%81%E5%8F%AC%E5%BC%80%E4%B8%AD%E5%A4%AE%E6%B7%B1%E6%94%B9%E5%A7%94%E4%BC%9A%E8%AE%AE&sa=fyb_n_homepage&rsv_dl=fyb_n_homepage&from=super&cl=3&tn=baidutop10&fr=top1000&rsv_idx=2&hisfilter=1"
    headers = {
        # 'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0'
    }  # 请求头，模拟浏览器访问

    # Referer:
    # 也被称为Referer或Referrer，这个头部指示了发出请求的页面（即当前页面是通过哪个URL链接过来的）。它对于分析网站流量来源和构建反向链接（backlinks）等场景非常有用。出于隐私考虑，用户可以选择让浏览器不发送这个头部，或者通过某些方法修改或伪造它。
    # Sec-Ch-Ua:
    # 这是一个与浏览器品牌和用户代理字符串相关的头部，它是"Client Hints"的一部分，用于向服务器提供关于用户代理的更详细和有用的信息，同时保护用户隐私。Sec-Ch-Ua通常包含浏览器品牌的名称和版本信息，但以一种加密和令牌化的形式发送，以保护用户的隐私。
    # Sec-Ch-Ua-Mobile:
    # 这也是"Client Hints"的一部分。这个头部用于指示用户代理是否是移动设备。如果请求来自移动设备，这个头部通常会包含一个特定的值（例如"1"），否则它可能不包含或包含其他值。
    # Sec-Ch-Ua-Platform:
    # 同样作为"Client Hints"的一部分，这个头部提供了关于用户代理运行平台的信息，例如操作系统类型和版本。和Sec-Ch-Ua一样，这些信息也是以加密和令牌化的形式发送的，以保护用户隐私。
    # User-Agent:
    # 这是一个传统的HTTP头部，用于标识发出请求的客户端类型，例如浏览器类型和版本、操作系统等。这个头部在过去被广泛用于网站的内容协商（content negotiation）和适应性设计（adaptive design），但由于其可以被伪造和滥用，现在许多现代网站和服务更倾向于依赖其他机制（如"Client Hints"）来获取更精确和隐私友好的用户代理信息。

    response = requests.get(url, headers=headers)  # 发送GET请求
    response.encoding = 'utf-8'  # 设置编码方式

    # soup = BeautifulSoup(response.text, 'html.parser')  # 解析网页内容
    soup = BeautifulSoup(response.text, 'lxml')  # 解析网页内容

    # BeautifulSoup支持的解析器主要有四种：
    # html.parser：这是Python内置的HTML解析器，不需要额外安装。它通常是BeautifulSoup的默认解析器，但在处理某些复杂的HTML文档时可能表现不佳。
    # lxml：lxml是一个高性能的解析库，需要额外安装。它通常比内置的html.parser解析器更快，而且能够处理一些复杂的HTML结构。要使用lxml解析器，你可以使用pip命令进行安装：pip install lxml。
    # lxml-xml：这是lxml库的另一个模式，用于解析XML文档。
    # html5lib：html5lib解析器是一个纯Python实现的解析器，它以与浏览器一致的方式解析HTML。它通常需要额外安装，可以使用pip命令进行安装：pip install html5lib。尽管它比内置的html.parser解析器慢，但它具有更好的容错性，可以处理包含错误或不完整的HTML。

    # 使用BeautifulSoup的方法提取需要的数据
    title = soup.title.string  # 提取网页标题
    links = [a['href'] for a in soup.find_all('a', href=True)]  # 提取网页中所有链接

    obj = [a.get('target') for a in soup.find_all('a', href=True)]
    obj = [a['target'] for a in soup.find_all('a', href=True) if 'target' in a.attrs]
    obj = [target for a in soup.find_all('a', href=True) if (target := a.get('target')) is not None]

    print('Title:', title)
    # print('Links:', links)
    # print('a=:', soup.find_all('a', href=True)[1])
    # print('b=:', soup.find_all('a', href=True))
    print('b=:', soup.find_all('a', title=True))
    print('c=:', soup.find_all('input'))

    wb = Workbook()
    # ws = wb.active
    # sheet_to_remove = wb["MySheet"]
    # wb.remove(sheet_to_remove)
    ws = wb.create_sheet("MySheet")
    sheet_to_remove = wb["Sheet"]
    wb.remove(sheet_to_remove)

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # sheet = wb["MySheet"]

    ws['A1'].fill = fill

    # i = 1
    # for a in soup.find_all('a'):
    #     print("aaa:",a)
    #     print("ppp:", a.get('value'))
    #     print("vvv:", a.find('span'))
    #     print("1", "2", "3", "3", "3", "3", "3","1", "2", "3", "3", "3", "3", "3")
    #     print("QQQ:",a.text)
    #     ws['A'+str(i)] = a.text
    #     # print(a['style'])
    #     # print("---------------------------------------------------------------")
    #     # print(re.compile(r'float:\s*([^;]+)').search("float:left;").group(0))
    #     # print(re.search(r'float:\s*([^;]+)', a['style']))
    #     # print(re.search(r'float:\s*([^;]+)', a['style']).group(1))
    #     # print(a.parent)
    #     i += 1
    #     if i == 0:
    #         break
    # else:
    #     print("OK")
    #     # print(a)
    # print('c=:', obj)
    # print(soup.text)

    # i = 1
    # for a in soup.find_all(['h1', 'h2', 'h3', 'li', 'p']):
    #     if not a.find('a'):
    #         ws['A'+str(i)] = a.text
    #         i += 1
    #     if i == 0:
    #         break
    # else:
    #     print("OK")

    i = 1
    aaa = lambda tag: tag.name != 'li' or len(tag.find_all('a')) == 0
    li_tags_without_a = soup.find_all(['h1', 'h2', 'h3', 'li', 'p'])
    for a in li_tags_without_a:
        if aaa(a):
            ws['A' + str(i)] = a.text
            i += 1
        else:
            pass

        if i == 0:
            break
    else:
        print("OK")

    # # 检查soup对象是否为None
    # if soup is not None:
    #     # 查找所有的<h1>, <h2>, <h3>, <li>, <p>标签，但是不包括那些包含<a>标签的<li>标签
    #     for a in soup.find_all(['h1', 'h2', 'h3', 'li', 'p'], lambda tag: tag.name != 'li' or len(tag.find_all('a')) == 0):
    #         print(a.name, a.text)
    # else:
    #     print("soup对象是None，请确保已经正确解析了HTML文档。")

    wb.save('sample.xlsx')

    print("---------------------------------------------------------------")
    # session = requests.Session()
    # # 第一个请求，获取cookie
    # response1 = session.get('http://www.example.com/login')
    #
    # # 后续的请求，session会自动附带之前获取的cookie
    # response2 = session.get('http://www.example.com/profile')
    #
    # # print(response2.text)
    # print(session.get(url).text)

    print("Hello", end="\n999")
    print("World!")  # 这将在同一行打印，因为前一个print的end参数是空格
    print(1, 2, 3, sep="-")  # 输出：1-2-3

    # print("Starting...", flush=True)
    print("Starting...")
    time.sleep(1)
    print("Finished!")

    arr = numpy.array([1, 2, 3, 4, 5])
    print(arr)
    print(numpy.__version__)

    # openpyxl
    # openpyxl是一个专门用于读写Excel 2010 xlsx/xlsm/xltx/xltm文件的库。它提供了对Excel工作簿、工作表、单元格等的底层访问，并允许你执行复杂的操作，如添加、删除和修改工作表，单元格样式，公式等。由于它的底层访问特性，openpyxl适合进行精细的Excel文件操作，特别是当你需要处理Excel文件的复杂结构或样式时。
    #
    # pandas
    # pandas是一个强大的数据分析库，它提供了DataFrame这一核心数据结构，用于高效地处理和分析结构化数据。pandas的read_excel和to_excel方法使得读写Excel文件变得非常简便。尽管pandas不是专为Excel设计的，但由于其强大的数据处理能力，它成为了处理Excel数据的常用工具。
    #
    # pandas适合进行大规模的数据处理和分析，它提供了丰富的数据清洗、转换、聚合等功能。然而，pandas在处理Excel文件的样式和复杂结构方面不如openpyxl灵活。
    #
    # 总结
    # openpyxl更专注于Excel文件的底层操作和样式处理。
    # pandas更适合进行大规模的数据处理和分析。


if __name__ == "__main__":
    main()
