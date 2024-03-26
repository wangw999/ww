import base64
import os
from datetime import datetime

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from pymongo.errors import OperationFailure, DuplicateKeyError
# import pandas as pd
from openpyxl import load_workbook
from django.core.files.storage import FileSystemStorage

from myapp.models import MyModel


# from myapp.myModel import MyModel

@csrf_exempt  # 如果你在前端处理了 CSRF，你可以使用 @csrf_exempt 来免除 CSRF 保护
@require_http_methods(["GET", "POST"])  # 限制只接受 GET 或 POST 请求
def downloadFile(request):
    if request.method == 'POST':
        # # path = "C:\Users\ZZ06NM672\Downloads\Mongodb.xlsx"
        # # 定义一个字符串
        # s = "Hello, World!"
        # # 使用encode方法将字符串转换为UTF-8编码的字节字符串
        # bytes_s = s.encode('utf-8')
        # # 将字节串转换为Base64编码的字符串
        # file_base64 = base64.b64encode(bytes_s).decode('utf-8')

        file_path = 'C:/Users/ZZ06NM672/Downloads/Mongodb.xlsx'  # 替换为你的文件路径
        base64_encoded_file = file_to_base64(file_path)

        # 获取当前时间，并格式化为字符串
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d%H%M%S")

        # 创建带有时间戳的文件名
        filename_with_timestamp = f"Mongodb_{timestamp}.xlsx"

        return JsonResponse({'fileName': filename_with_timestamp, 'fileBase64': base64_encoded_file})

    elif request.method == 'GET':
        return JsonResponse({'data': '这是从服务器获取的数据'})


def file_to_base64(file_path):
    with open(file_path, 'rb') as file:
        # 读取文件内容
        file_content = file.read()

        # 将文件内容转换为base64编码的字节流
        base64_bytes = base64.b64encode(file_content)

        # 如果需要字符串格式，可以解码为utf-8
        base64_string = base64_bytes.decode('utf-8')

        return base64_string


@csrf_exempt  # 如果你不打算使用CSRF保护，或者你已经配置了CSRF中间件以允许POST请求
@require_http_methods(["POST"])
def upload_file_view(request):
    if request.FILES.get('myFile'):
        excel_file = request.FILES['myFile']
        # 在这里处理文件，例如保存到硬盘或数据库中
        # ...

        # # 使用pandas读取Excel文件
        # df = pd.read_excel(excel_file)

        # 获取当前时间，并格式化为字符串
        now = datetime.now()
        timestamp = now.strftime("%Y%m%d%H%M%S")

        # 设置文件存储路径
        fs = FileSystemStorage()
        # 生成新的文件名，这里使用了UUID以确保唯一性
        new_filename = os.path.splitext(excel_file.name)[0] + "_" +timestamp + "." + excel_file.name.split('.')[-1]
        filename = fs.save(new_filename, excel_file)
        uploaded_file_url = fs.url(filename + timestamp)

        # 使用openpyxl加载上传的文件
        workbook = load_workbook(excel_file)
        sheet = workbook.active  # 获取活动工作表

        # 读取工作表数据并返回给前端
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
            # print(row)

            # 返回一个JSON响应
        return JsonResponse({'status': 'success', 'message': '文件已上传' + uploaded_file_url})
    else:
        return JsonResponse({'status': 'error', 'message': '没有接收到文件'}, status=400)


@csrf_exempt  # 如果你在前端处理了 CSRF，你可以使用 @csrf_exempt 来免除 CSRF 保护
@require_http_methods(["GET", "POST"])  # 限制只接受 GET 或 POST 请求
def my_ajax_view(request):
    if request.method == 'POST':
        # 返回JSON
        # # 处理 POST 请求，比如获取表单数据
        # data = request.POST.get('data')
        # # 处理数据...
        # return JsonResponse({'success': True, 'message': '数据处理成功'})

        # mongoDbTest Start
        # 创建并保存对象
        obj = MyModel(field1='111', field2='12')
        obj.save()

        # try:
        #     # 尝试执行一些数据库操作
        #     obj = MyModel(field1='value1', field2='12')
        #     obj.save()
        # except DuplicateKeyError as e:
        #     # 处理重复键错误
        #     print(f"A duplicate key error occurred: {e}")
        # except OperationFailure as e:
        #     # 处理其他操作失败的情况
        #     print(f"An operation failure occurred: {e}")
        # except Exception as e:
        #     # 处理其他未特定捕获的异常
        #     print(f"An unexpected error occurred: {e}")
        # finally:
        #     # 无论是否发生异常，都可以在这里执行清理操作
        #     pass

        # # 查询对象
        # objs = MyModel.objects.all()
        # for obj in objs:
        #     print(obj.field1, obj.field2)

        # # 根据条件查询
        # objs_filtered = MyModel.objects.filter(field1='value1')
        #
        # # 更新对象
        # obj.field1 = 'new_value'
        # obj.save()

        # mongoDbTest End

        # --------------------------------------------------------------
        # # 返回HTML
        # context = {
        #     'name': 'John Doe',
        #     'greeting': 'Hello there!',
        #     'items': ['apple', 'banana', 'cherry'],
        # }
        #
        # # 使用 'my_template.html' 模板和上下文数据渲染页面
        # html_content = render_to_string('my_template.html', context)
        # return HttpResponse(html_content, content_type='text/html')
        # --------------------------------------------------------------

        # --------------------------------------------------------------
        aa = '''<!DOCTYPE html>  
                <html lang="zh-CN">  
                <head>  
                    <meta charset="UTF-8">  
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">  
                    <title>简单的HTML示例</title>  
                </head>  
                <body>  
                    <h1>欢迎来到我的网页</h1>  
                    <p>这是一个简单的HTML示例页面。</p>  
                    <p>你可以在这里添加更多的内容，比如段落、图片、链接等。</p>  
                    <a href="https://www.example.com">点击这里访问示例网站</a>  
                </body>  
                </html>'''

        return HttpResponse(aa, content_type='text/html')
        # --------------------------------------------------------------

        # if html_content is not None:
        #     return HttpResponse(html_content, content_type='text/html')
        # else:
        #     return HttpResponse("", content_type='text/html')

    elif request.method == 'GET':
        # 处理 GET 请求，比如获取数据
        return JsonResponse({'data': '这是从服务器获取的数据'})

# Step Over (F8): 执行当前行，如果当前行是一个函数调用，则执行整个函数并停在下一行。
# Step Into (F7): 如果当前行是一个函数调用，则进入该函数并在第一行暂停；如果不是函数调用，则等同于 Step Over。
# Step Out (Shift + F8): 执行到当前函数或方法的返回语句，并停在调用该函数或方法的下一行。
